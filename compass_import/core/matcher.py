"""
core/matcher.py
──────────────────────────────────────────────────────────────
Compass · Consumer Voice — Import Pipeline
Module 3 — Matching catégories : détection, export XLS, validation, propagation.

Flux d'utilisation typique :
    products   = get_unmatched_products(conn)
    xls_bytes  = export_matching_xls(products, referentiel)
    # … opérateur remplit le XLS dans Excel …
    result     = validate_matching_xls(xls_bytes, referentiel, original_names)
    stats      = apply_matching(conn, result["valid"])
"""

import io
import logging
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from psycopg2.extras import execute_values

if sys.version_info >= (3, 11):
    import tomllib
else:
    import tomli as tomllib

logger = logging.getLogger(__name__)

_CONFIG_PATH = Path(__file__).parent.parent / "config.toml"

# Style constants (from config.toml [matching] — hard-coded here as fallback)
_HEADER_FILL   = "1F6ED4"  # Compass blue
_HEADER_FONT   = "FFFFFF"  # white text
_EDITABLE_FILL = "FFFBEB"  # yellow-ish — editable cells
_LOCKED_FILL   = "F3F4F6"  # light gray — locked cells
_REF_FILL      = "E8F0FE"  # light blue — Référentiel headers
_VALID_PHOTO   = {"true", "false"}


def _load_config() -> dict:
    try:
        with open(_CONFIG_PATH, "rb") as f:
            return tomllib.load(f)
    except FileNotFoundError:
        return {}


def _sanitize_excel_name(name: str) -> str:
    """
    Convertit un nom de catégorie en identifiant Excel valide pour
    les named ranges (INDIRECT). Remplace les espaces et caractères
    spéciaux par des underscores.
    """
    result = ""
    for ch in name:
        result += ch if ch.isalnum() else "_"
    # Les named ranges Excel ne peuvent pas commencer par un chiffre
    if result and result[0].isdigit():
        result = "_" + result
    return result


# ─── get_unmatched_products ───────────────────────────────────────────────────

def get_unmatched_products(conn) -> list[dict]:
    """
    Retourne les produits présents dans ``verbatims`` mais absents de
    ``categories_mapping`` (jointure sur brand + product_name).

    Trie par nombre de verbatims décroissant pour aider l'opérateur à
    prioriser les produits les plus impactants.

    Args:
        conn: Connexion psycopg2 active.

    Returns:
        Liste de dicts avec clés ``brand``, ``product_name``, ``nb_verbatims``.
    """
    query = """
        SELECT v.brand, v.product_name, COUNT(*) AS nb_verbatims
          FROM verbatims v
          LEFT JOIN categories_mapping cm
                 ON v.brand = cm.brand AND v.product_name = cm.product_name
         WHERE cm.key_brandxpdt IS NULL
         GROUP BY v.brand, v.product_name
         ORDER BY nb_verbatims DESC
    """
    with conn.cursor() as cur:
        cur.execute(query)
        cols = [desc[0] for desc in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


# ─── export_matching_xls ─────────────────────────────────────────────────────

def export_matching_xls(products: list[dict], referentiel: dict) -> bytes:
    """
    Génère un fichier XLS (.xlsx) pour le matching des catégories.

    Structure :
    - Onglet 1 "Matching" : une ligne par produit à matcher.
      Colonnes A-B verrouillées (product_name, nb_verbatims).
      Colonnes C-E éditables avec menus déroulants (categorie, sous-catégorie, photo).
    - Onglet 2 "Référentiel" : toutes les combinaisons valides, lecture seule.
      Sert aussi de source pour les dropdowns de l'onglet Matching.

    Les dropdowns de sous-catégorie sont dépendants de la catégorie sélectionnée
    via des named ranges et la formule INDIRECT(SUBSTITUTE(Cn," ","_")).

    Args:
        products: Liste de dicts ``{product_name, nb_verbatims}`` — sortie
                  de ``get_unmatched_products()``.
        referentiel: Dict ``{categorie: [sous_categories]}`` — sortie de
                     ``load_referentiel()``.

    Returns:
        Contenu binaire du fichier .xlsx.
    """
    config   = _load_config().get("matching", {})
    hdr_fill = config.get("header_fill_color", _HEADER_FILL)
    hdr_font = config.get("header_font_color", _HEADER_FONT)
    edt_fill = config.get("editable_fill_color", _EDITABLE_FILL)
    lck_fill = config.get("locked_fill_color", _LOCKED_FILL)

    wb = Workbook()

    # ──────────────────────────────────────────────────────────────────────────
    # Onglet 2 — Référentiel (créé en premier pour les named ranges)
    # ──────────────────────────────────────────────────────────────────────────
    ws_ref = wb.create_sheet(title="Référentiel")

    categories  = sorted(referentiel.keys())
    all_sous    = [s for cat in categories for s in referentiel[cat]]
    n_cats      = len(categories)
    n_sous_all  = len(all_sous)

    ref_hdr_fill = PatternFill("solid", fgColor=_REF_FILL)
    ref_hdr_font = Font(bold=True, color=_HEADER_FILL, name="Calibri", size=10)

    # Colonne A : liste des catégories (source du dropdown Catégorie)
    ws_ref.cell(row=1, column=1, value="Catégorie").font = ref_hdr_font
    ws_ref.cell(row=1, column=1).fill = ref_hdr_fill
    for i, cat in enumerate(categories, start=2):
        ws_ref.cell(row=i, column=1, value=cat)

    # Colonne B : liste plate de toutes les sous-catégories (référence lisible)
    ws_ref.cell(row=1, column=2, value="Sous-catégorie").font = ref_hdr_font
    ws_ref.cell(row=1, column=2).fill = ref_hdr_fill
    for i, sous in enumerate(all_sous, start=2):
        ws_ref.cell(row=i, column=2, value=sous)

    # Colonnes D+ : une colonne par catégorie avec ses sous-catégories
    # → sources des named ranges pour les dropdowns dépendants
    col_offset = 4  # colonne D
    for cat_idx, cat in enumerate(categories):
        col_idx   = col_offset + cat_idx
        col_letter = get_column_letter(col_idx)
        sous_list  = referentiel[cat]

        hdr_cell = ws_ref.cell(row=1, column=col_idx, value=cat)
        hdr_cell.font = ref_hdr_font
        hdr_cell.fill = ref_hdr_fill

        for row_offset, sous in enumerate(sous_list, start=2):
            ws_ref.cell(row=row_offset, column=col_idx, value=sous)

        row_start = 2
        row_end   = 1 + len(sous_list)

        # Named range : "Body_Care" → Référentiel!$D$2:$D$3
        safe_name   = _sanitize_excel_name(cat)
        ref_formula = f"'Référentiel'!${col_letter}${row_start}:${col_letter}${row_end}"
        wb.defined_names[safe_name] = DefinedName(
            name=safe_name, attr_text=ref_formula
        )

    # Largeurs colonnes Référentiel
    ws_ref.column_dimensions["A"].width = 22
    ws_ref.column_dimensions["B"].width = 40
    # Protéger le Référentiel en lecture seule
    ws_ref.protection.sheet = True
    ws_ref.protection.enable()
    ws_ref.freeze_panes = "A2"

    # ──────────────────────────────────────────────────────────────────────────
    # Onglet 1 — Matching
    # ──────────────────────────────────────────────────────────────────────────
    ws_match = wb.active
    ws_match.title = "Matching"

    match_hdr_fill = PatternFill("solid", fgColor=hdr_fill)
    match_hdr_font = Font(bold=True, color=hdr_font, name="Calibri", size=11)
    edt_bg         = PatternFill("solid", fgColor=edt_fill)
    lck_bg         = PatternFill("solid", fgColor=lck_fill)
    align_center   = Alignment(horizontal="center", vertical="center")
    align_left     = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Headers ───────────────────────────────────────────────────────────────
    # Colonnes A-D verrouillées, colonnes E-G éditables
    _COLS = [
        ("key_brandxpdt",         False),  # A — locked  (brand || product_name)
        ("brand",                 False),  # B — locked
        ("product_name",          False),  # C — locked
        ("nb_verbatims",          False),  # D — locked
        ("categorie_interne",     True),   # E — editable
        ("sous_categorie_interne",True),   # F — editable
        ("photo",                 True),   # G — editable
    ]
    for col_idx, (label, _editable) in enumerate(_COLS, start=1):
        cell = ws_match.cell(row=1, column=col_idx, value=label)
        cell.fill      = match_hdr_fill
        cell.font      = match_hdr_font
        cell.alignment = align_center
        cell.protection = Protection(locked=True)

    # ── Data rows ─────────────────────────────────────────────────────────────
    n_rows = len(products)
    for row_idx, product in enumerate(products, start=2):
        brand_val = product.get("brand", "")
        pname_val = product.get("product_name", "")

        # A — key_brandxpdt (locked)
        c = ws_match.cell(row=row_idx, column=1, value=brand_val + pname_val)
        c.fill = lck_bg; c.alignment = align_left; c.protection = Protection(locked=True)

        # B — brand (locked)
        c = ws_match.cell(row=row_idx, column=2, value=brand_val)
        c.fill = lck_bg; c.alignment = align_left; c.protection = Protection(locked=True)

        # C — product_name (locked)
        c = ws_match.cell(row=row_idx, column=3, value=pname_val)
        c.fill = lck_bg; c.alignment = align_left; c.protection = Protection(locked=True)

        # D — nb_verbatims (locked)
        c = ws_match.cell(row=row_idx, column=4, value=product.get("nb_verbatims", 0))
        c.fill = lck_bg; c.alignment = align_center; c.protection = Protection(locked=True)

        # E — categorie_interne (editable)
        c = ws_match.cell(row=row_idx, column=5)
        c.fill = edt_bg; c.alignment = align_center; c.protection = Protection(locked=False)

        # F — sous_categorie_interne (editable)
        c = ws_match.cell(row=row_idx, column=6)
        c.fill = edt_bg; c.alignment = align_center; c.protection = Protection(locked=False)

        # G — photo (editable)
        c = ws_match.cell(row=row_idx, column=7)
        c.fill = edt_bg; c.alignment = align_center; c.protection = Protection(locked=False)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws_match.column_dimensions["A"].width = 50   # key_brandxpdt
    ws_match.column_dimensions["B"].width = 22   # brand
    ws_match.column_dimensions["C"].width = 42   # product_name
    ws_match.column_dimensions["D"].width = 14   # nb_verbatims
    ws_match.column_dimensions["E"].width = 26   # categorie_interne
    ws_match.column_dimensions["F"].width = 38   # sous_categorie_interne
    ws_match.column_dimensions["G"].width = 10   # photo

    # ── Data validations (only if there are rows) ─────────────────────────────
    if n_rows > 0:
        last_row = n_rows + 1

        # Catégorie — dropdown depuis Référentiel!$A$2:$A${n_cats+1}
        cat_dv = DataValidation(
            type="list",
            formula1=f"'Référentiel'!$A$2:$A${n_cats + 1}",
            allow_blank=True,
            showDropDown=False,
            errorTitle="Catégorie invalide",
            error="Choisissez une catégorie dans la liste.",
            errorStyle="stop",
        )
        cat_dv.add(f"E2:E{last_row}")
        ws_match.add_data_validation(cat_dv)

        # Sous-catégorie — dropdown dépendant via INDIRECT + named ranges
        # La formule INDIRECT(SUBSTITUTE(E2," ","_")) référence le named range
        # correspondant à la catégorie sélectionnée en colonne E.
        sous_dv = DataValidation(
            type="list",
            formula1='INDIRECT(SUBSTITUTE(E2," ","_"))',
            allow_blank=True,
            showDropDown=False,
        )
        sous_dv.add(f"F2:F{last_row}")
        ws_match.add_data_validation(sous_dv)

        # Photo — dropdown true / false
        photo_dv = DataValidation(
            type="list",
            formula1='"true,false"',
            allow_blank=False,
            showDropDown=False,
            errorTitle="Valeur photo invalide",
            error="Choisissez 'true' ou 'false'.",
            errorStyle="stop",
        )
        photo_dv.add(f"G2:G{last_row}")
        ws_match.add_data_validation(photo_dv)

    # ── Sheet protection — seules C-E sont déverrouillées ────────────────────
    ws_match.protection.sheet = True
    ws_match.protection.enable()
    ws_match.freeze_panes = "A2"

    # ── Serialize ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── validate_matching_xls ───────────────────────────────────────────────────

def validate_matching_xls(
    file_bytes: bytes,
    referentiel: dict,
    original_keys: set[str] | None = None,
) -> dict:
    """
    Valide le fichier XLS de matching complété par l'opérateur.

    Structure attendue (v1.3) :
      A=key_brandxpdt  B=brand  C=product_name  D=nb_verbatims (locked)
      E=categorie_interne  F=sous_categorie_interne  G=photo (editable)

    Contrôles ligne par ligne :
    - ``key_brandxpdt`` vide → rejet.
    - ``key_brandxpdt`` modifié (si ``original_keys`` fourni) → rejet.
    - ``categorie_interne`` vide → rejet.
    - ``sous_categorie_interne`` vide → rejet.
    - ``photo`` ni "true" ni "false" → rejet.
    - Combinaison catégorie / sous-catégorie hors référentiel → rejet.

    Args:
        file_bytes: Contenu binaire du fichier .xlsx.
        referentiel: Dict ``{categorie: [sous_categories]}`` — référentiel fermé.
        original_keys: Ensemble des ``key_brandxpdt`` attendus (optionnel).
                       Si fourni, détecte les clés inconnues ou modifiées.

    Returns:
        Dict ``{"valid": list[dict], "errors": list[dict]}``.
        Chaque entrée d'erreur contient : ``ligne``, ``key_brandxpdt``,
        ``colonne``, ``valeur``, ``raison``.
        Chaque entrée valide contient : ``key_brandxpdt``, ``brand``,
        ``product_name``, ``categorie_interne``, ``sous_categorie_interne``,
        ``photo`` (bool).

    Raises:
        ValueError: Si le fichier est illisible ou ne contient pas l'onglet
                    "Matching".
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier XLS : {exc}") from exc

    if "Matching" not in wb.sheetnames:
        raise ValueError(
            "L'onglet 'Matching' est introuvable dans le fichier.\n"
            "Vérifiez que vous avez bien importé le fichier exporté par l'application."
        )

    ws = wb["Matching"]
    valid_rows:  list[dict] = []
    error_rows:  list[dict] = []

    def _err(row_idx, key, colonne, valeur, raison):
        error_rows.append({
            "ligne":          row_idx,
            "key_brandxpdt":  key,
            "colonne":        colonne,
            "valeur":         str(valeur),
            "raison":         raison,
        })

    def _cell_str(val) -> str:
        if val is None:
            return ""
        if isinstance(val, bool):
            return "true" if val else "false"
        return str(val).strip()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Ignorer les lignes entièrement vides
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        # A=0  B=1  C=2  D=3  E=4  F=5  G=6
        key_val        = _cell_str(row[0] if len(row) > 0 else None)
        brand_val      = _cell_str(row[1] if len(row) > 1 else None)
        product_val    = _cell_str(row[2] if len(row) > 2 else None)
        categorie      = _cell_str(row[4] if len(row) > 4 else None)
        sous_categorie = _cell_str(row[5] if len(row) > 5 else None)
        photo_raw      = _cell_str(row[6] if len(row) > 6 else None).lower()

        row_errors: list[tuple] = []

        # key_brandxpdt vide
        if not key_val:
            _err(row_idx, "", "key_brandxpdt", "", "key_brandxpdt vide — ligne ignorée")
            continue

        # key_brandxpdt modifié / inconnu
        if original_keys is not None and key_val not in original_keys:
            row_errors.append(("key_brandxpdt", key_val,
                               f"Clé inconnue ou modifiée : '{key_val}'"))

        # catégorie vide
        if not categorie:
            row_errors.append(("categorie_interne", "", "Catégorie vide"))

        # sous-catégorie vide
        if not sous_categorie:
            row_errors.append(("sous_categorie_interne", "", "Sous-catégorie vide"))

        # photo invalide
        if photo_raw not in _VALID_PHOTO:
            row_errors.append(("photo", photo_raw,
                               f"Valeur photo invalide : '{photo_raw}' (attendu : true ou false)"))

        # combinaison hors référentiel (seulement si les deux champs sont remplis)
        if categorie and sous_categorie:
            if sous_categorie not in referentiel.get(categorie, []):
                row_errors.append((
                    "sous_categorie_interne",
                    f"{categorie} / {sous_categorie}",
                    f"Combinaison hors référentiel : '{categorie}' / '{sous_categorie}'",
                ))

        if row_errors:
            for colonne, valeur, raison in row_errors:
                _err(row_idx, key_val, colonne, valeur, raison)
        else:
            valid_rows.append({
                "key_brandxpdt":         key_val,
                "brand":                 brand_val,
                "product_name":          product_val,
                "categorie_interne":     categorie,
                "sous_categorie_interne": sous_categorie,
                "photo":                 photo_raw == "true",
            })

    return {"valid": valid_rows, "errors": error_rows}


# ─── apply_matching ───────────────────────────────────────────────────────────

def apply_matching(conn, valid_rows: list[dict]) -> dict:
    """
    Applique le matching : UPSERT dans ``categories_mapping`` puis UPDATE
    en cascade sur ``verbatims``.

    La propagation est totale : pour chaque produit, **tous** les verbatims
    existants (y compris les imports antérieurs) reçoivent les nouvelles valeurs.
    Si une catégorie est corrigée, la correction se propage sur l'intégralité
    de l'historique du produit.

    Args:
        conn: Connexion psycopg2 active (``autocommit=False``).
        valid_rows: Liste de dicts validés — sortie de ``validate_matching_xls``.
                    Clés attendues : ``product_name``, ``categorie_interne``,
                    ``sous_categorie_interne``, ``photo`` (bool).

    Returns:
        Dict ``{"products_matched": int, "verbatims_updated": int}``.
    """
    if not valid_rows:
        return {"products_matched": 0, "verbatims_updated": 0}

    # ── UPSERT categories_mapping ──────────────────────────────────────────────
    upsert_sql = """
        INSERT INTO categories_mapping
            (key_brandxpdt, brand, product_name,
             categorie_interne, sous_categorie_interne, photo)
        VALUES %s
        ON CONFLICT (key_brandxpdt) DO UPDATE SET
            categorie_interne      = EXCLUDED.categorie_interne,
            sous_categorie_interne = EXCLUDED.sous_categorie_interne,
            photo                  = EXCLUDED.photo,
            matched_at             = NOW()
    """
    upsert_values = [
        (
            row["key_brandxpdt"],
            row["brand"],
            row["product_name"],
            row["categorie_interne"],
            row["sous_categorie_interne"],
            row["photo"],
        )
        for row in valid_rows
    ]
    with conn.cursor() as cur:
        execute_values(cur, upsert_sql, upsert_values)
    logger.info("categories_mapping : %d produits upsertés", len(valid_rows))

    # ── UPDATE verbatims en cascade ────────────────────────────────────────────
    # Jointure sur (brand, product_name) pour propager sur tous les verbatims
    # du produit, quelle que soit la date d'import.
    update_sql = """
        UPDATE verbatims AS v SET
            categorie_interne      = m.categorie::varchar,
            sous_categorie_interne = m.sous_cat::varchar,
            photo                  = m.photo::boolean
        FROM (VALUES %s) AS m(brand, product_name, categorie, sous_cat, photo)
        WHERE v.brand = m.brand AND v.product_name = m.product_name
    """
    update_values = [
        (
            row["brand"],
            row["product_name"],
            row["categorie_interne"],
            row["sous_categorie_interne"],
            str(row["photo"]).lower(),
        )
        for row in valid_rows
    ]
    with conn.cursor() as cur:
        execute_values(cur, update_sql, update_values)
        verbatims_updated = cur.rowcount

    conn.commit()
    logger.info(
        "verbatims mis à jour : %d lignes sur %d produits",
        verbatims_updated, len(valid_rows),
    )

    return {
        "products_matched":  len(valid_rows),
        "verbatims_updated": verbatims_updated,
    }
